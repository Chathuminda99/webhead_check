import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import argparse
from tqdm import tqdm

# ASCII Art Title
print("""
                         /$$       /$$                                 /$$                 /$$                           /$$      
                        | $$      | $$                                | $$                | $$                          | $$      
 /$$  /$$  /$$  /$$$$$$ | $$$$$$$ | $$$$$$$   /$$$$$$   /$$$$$$   /$$$$$$$        /$$$$$$$| $$$$$$$   /$$$$$$   /$$$$$$$| $$   /$$
| $$ | $$ | $$ /$$__  $$| $$__  $$| $$__  $$ /$$__  $$ |____  $$ /$$__  $$       /$$_____/| $$__  $$ /$$__  $$ /$$_____/| $$  /$$/
| $$ | $$ | $$| $$$$$$$$| $$  \ $$| $$  \ $$| $$$$$$$$  /$$$$$$$| $$  | $$      | $$      | $$  \ $$| $$$$$$$$| $$      | $$$$$$/ 
| $$ | $$ | $$| $$_____/| $$  | $$| $$  | $$| $$_____/ /$$__  $$| $$  | $$      | $$      | $$  | $$| $$_____/| $$      | $$_  $$ 
|  $$$$$/$$$$/|  $$$$$$$| $$$$$$$/| $$  | $$|  $$$$$$$|  $$$$$$$|  $$$$$$$      |  $$$$$$$| $$  | $$|  $$$$$$$|  $$$$$$$| $$ \  $$
 \_____\___/  \_______/|_______/ |__/  |__/ \_______/ \_______/ \_______//$$$$$$\_______/|__/  |__/ \_______/ \_______/|__/  \__/
                                                                         |______/                                                 
                                                                                                                                  
                                                                                                                                  
""")

# List of security headers to check
SECURITY_HEADERS = [
    "X-Frame-Options",
    "X-Content-Type-Options",
    "Strict-Transport-Security",
    "Content-Security-Policy",
    "Referrer-Policy",
    "Permissions-Policy",
    "Cross-Origin-Embedder-Policy",
    "Cross-Origin-Resource-Policy",
    "Cross-Origin-Opener-Policy"
]

# Function to fetch headers for a URL
def fetch_headers(url):
    try:
        response = requests.head(url, timeout=10)
        return response.headers
    except requests.RequestException as e:
        print(f"Error fetching headers for {url}: {e}")
        return {}

# Read URLs from text file
def read_urls_from_file(file_path):
    with open(file_path, 'r') as file:
        urls = [line.strip() for line in file.readlines()]
    return urls

# Generate Excel sheet with header information
def generate_excel(urls, output_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Security Headers Report"

    # Add headers to the Excel sheet
    headers = ["URL"] + SECURITY_HEADERS
    sheet.append(headers)

    # Styles for cells
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)

    # Style header row
    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = bold_font

    # Fetch headers for each URL with progress view
    for url in tqdm(urls, desc="Processing URLs", unit="url"):
        fetched_headers = fetch_headers(url)
        row = [url]
        for header in SECURITY_HEADERS:
            status = "Header Implemented" if header in fetched_headers else "Header Not Implemented"
            row.append(status)
        sheet.append(row)

        # Apply color coding
        current_row = sheet.max_row
        for col_index, header in enumerate(SECURITY_HEADERS, start=2):
            cell = sheet.cell(row=current_row, column=col_index)
            if cell.value == "Header Implemented":
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    # Save the workbook
    workbook.save(output_file)

# Display results for a single URL
def display_single_url_result(url):
    headers = fetch_headers(url)
    print(f"\nResults for URL: {url}\n")
    for header in SECURITY_HEADERS:
        status = "Header Implemented" if header in headers else "Header Not Implemented"
        print(f"{header}: {status}")

# Main script
def main():
    parser = argparse.ArgumentParser(description="Check security headers for a list of URLs and generate a report.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-u", "--url", help="Single URL to check.")
    group.add_argument("-l", "--list", help="Path to the file containing the list of URLs.")
    parser.add_argument("-o", "--output", required=False, default="security_headers_report.xlsx", help="Name of the output Excel file.")
    args = parser.parse_args()

    # Determine input source
    if args.url:
        display_single_url_result(args.url)
    elif args.list:
        urls = read_urls_from_file(args.list)

        output_file = args.output

        # Ensure output file has .xlsx extension
        if not output_file.endswith(".xlsx"):
            output_file += ".xlsx"

        # Generate Excel report
        generate_excel(urls, output_file)

        print(f"Security headers report generated successfully: {output_file}")

if __name__ == "__main__":
    main()
